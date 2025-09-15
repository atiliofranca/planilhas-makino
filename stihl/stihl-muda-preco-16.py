import pandas as pd
import tkinter as tk
from tkinter import filedialog
from screeninfo import get_monitors
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- FUNÇÕES AUXILIARES ---

def arredondamento_personalizado(numero):
    if not isinstance(numero, (int, float)):
        return numero
    numero = round(numero, 2)
    parte_decimal = round(numero % 1, 2)
    parte_inteira = int(numero)
    if parte_decimal == 0.00 or parte_decimal == 0.50:
        return float(numero)
    elif parte_decimal > 0.00 and parte_decimal < 0.50:
        return parte_inteira + 0.50
    elif parte_decimal > 0.50:
        return float(parte_inteira + 1)
    else:
        return float(numero)

def letra_para_indice(letra):
    letra = letra.upper()
    indice = 0
    for char in letra:
        indice = indice * 26 + (ord(char) - ord('A') + 1)
    return indice - 1

def centralizar_janela_raiz(root):
    try:
        monitor_principal = next(m for m in get_monitors() if m.is_primary)
    except StopIteration:
        monitor_principal = get_monitors()[0]
    largura_tela = monitor_principal.width
    altura_tela = monitor_principal.height
    offset_x = monitor_principal.x
    offset_y = monitor_principal.y
    pos_x = offset_x + (largura_tela // 2)
    pos_y = offset_y + (altura_tela // 2)
    root.geometry(f'+{pos_x}+{pos_y}')

# --- CONFIGURAÇÃO ---
nome_arquivo_saida_excel = 'stihl/new-stihl16.xlsx'
nome_arquivo_saida_autcom_csv = 'stihl/new-stihl16-autcom.csv'
COLUNA_BUSCA_NOVO_ARQ = 'B'
COLUNA_PRECO_BASE = 'E'

# --- LISTA DE EXCEÇÕES ---
REFERENCIAS_PARA_IGNORAR = [
    '1127-120-1620', '4112-713-4100', '4119-713-4100', '0000-881-9411',
    '0781-389-3004', '7030-319-0000', '7030-516-0000', '7030-319-0001',
    '7030-516-0002', '0781-389-3012', '7030-319-0002'
]

# --- SELEÇÃO DE ARQUIVOS COM POP-UP ---
root = tk.Tk()
centralizar_janela_raiz(root)
root.update()
root.withdraw()

print("Por favor, selecione os arquivos de entrada nas janelas pop-up...")
nome_arquivo_1 = filedialog.askopenfilename(
    title="Passo 1 de 2: Selecione o arquivo.csv do Autcom (Lista Base)",
    filetypes=(("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*"))
)
if not nome_arquivo_1: print("Seleção cancelada."); exit()

nome_arquivo_2 = filedialog.askopenfilename(
    title="Passo 2 de 2: Selecione a NOVA planilha da Stihl (Fonte de Preços)",
    filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
)
if not nome_arquivo_2: print("Seleção cancelada."); exit()

print(f"\nArquivo 1 (Base): {nome_arquivo_1}")
print(f"Arquivo 2 (Fonte de Preços): {nome_arquivo_2}\n")

# --- LEITURA E PRÉ-PROCESSAMENTO ---
try:
    df1 = pd.read_csv(nome_arquivo_1, sep=';', encoding='latin-1', decimal=',')
    
    print("Lendo a planilha de preços...")
    df_precos = pd.read_excel(nome_arquivo_2, header=None, engine='openpyxl')
    
    dados_precos = {}
    indice_busca = letra_para_indice(COLUNA_BUSCA_NOVO_ARQ)
    indice_preco_base = letra_para_indice(COLUNA_PRECO_BASE)
    for index, linha in df_precos.iterrows():
        referencia = linha.get(indice_busca)
        if pd.notna(referencia):
            dados_precos[referencia] = linha.get(indice_preco_base)
            
    print(f"{len(dados_precos)} referências de preço lidas com sucesso!\n")

except Exception as e:
    print(f"Ocorreu um erro ao ler ou processar os arquivos: {e}")
    exit()

# --- PROCESSAMENTO PRINCIPAL ---
indices_modificados = []
indices_ignorados = []
df1['Preço Real Stihl'] = pd.NA

colunas_venda = [
    'Novo Pr. Venda 1', 'Novo Pr. Venda 2', 'Novo Pr. Venda 3', 'Novo Pr. Venda 4',
    'Novo Pr. Venda 5', 'Novo Pr. Venda 6', 'Novo Pr. Venda 7'
]

for index, linha_df1 in df1.iterrows():
    valor_a_procurar = linha_df1['Referência']
    
    if valor_a_procurar in REFERENCIAS_PARA_IGNORAR:
        print(f"Aviso: Referência '{valor_a_procurar}' (linha {index + 2}) está na lista de exceções.")
        indices_ignorados.append(index)
        continue
    
    if pd.notna(valor_a_procurar) and valor_a_procurar in dados_precos:
        valor_base_encontrado = dados_precos[valor_a_procurar]
        
        try:
            valor_como_numero = float(str(valor_base_encontrado).replace(',', '.'))
            
            valor_venda_4 = arredondamento_personalizado(valor_como_numero)
            valor_venda_1 = arredondamento_personalizado(valor_venda_4 * 1.07)
            valor_venda_5 = arredondamento_personalizado(valor_venda_4 * 0.98)
            valor_venda_6 = valor_venda_1
            valor_venda_7 = valor_venda_1
            valor_venda_2 = arredondamento_personalizado(valor_venda_1 * 0.98)
            valor_venda_3 = arredondamento_personalizado(valor_venda_2 * 0.98)
            
            valor_final_compra = round(valor_venda_4 * 0.67, 2)
            valor_final_frete = round(valor_final_compra * 0.015, 2)
            
            print(f"Valor '{valor_a_procurar}' encontrado! Atualizando linha {index + 2}...")
            df1.loc[index, 'Novo Pr. Venda 1'] = valor_venda_1
            df1.loc[index, 'Novo Pr. Venda 2'] = valor_venda_2
            df1.loc[index, 'Novo Pr. Venda 3'] = valor_venda_3
            df1.loc[index, 'Novo Pr. Venda 4'] = valor_venda_4
            df1.loc[index, 'Novo Pr. Venda 5'] = valor_venda_5
            df1.loc[index, 'Novo Pr. Venda 6'] = valor_venda_6
            df1.loc[index, 'Novo Pr. Venda 7'] = valor_venda_7
            df1.loc[index, 'Novo Pr.Compra'] = valor_final_compra
            df1.loc[index, 'Novo Frete Entrada'] = valor_final_frete
            df1.loc[index, 'Preço Real Stihl'] = valor_base_encontrado
            
            indices_modificados.append(index)

        except (ValueError, TypeError):
            print(f"Aviso: Valor de preço base '{valor_base_encontrado}' não é numérico para a ref '{valor_a_procurar}'.")


# --- PREPARAÇÃO PARA SAÍDA ---
# MUDANÇA: Lista de colunas a serem removidas de TODOS os arquivos de saída
colunas_a_remover = [
    'Novo IPI Entrada', 
    'Novo Departamento', 
    'Desc. Departamento', 
    'Unnamed: 14', 
    'Unnamed: 15'
]
df1.drop(columns=colunas_a_remover, inplace=True, errors='ignore')


# --- GERAÇÃO ARQUIVO.XLSX ---
def destacar_celulas(linha):
    colunas_alteradas = colunas_venda + ['Novo Pr.Compra', 'Novo Frete Entrada']
    if linha.name in indices_ignorados:
        return ['background-color: #FFCDD2'] * len(linha)
    elif linha.name in indices_modificados:
        return ['background-color: yellow' if col in colunas_alteradas else '' for col in linha.index]
    else:
        return [''] * len(linha)

styled_df = df1.style.apply(destacar_celulas, axis=1)

colunas_para_formatar_excel = colunas_venda + ['Novo Pr.Compra', 'Novo Frete Entrada', 'Preço Real Stihl']
formatador_excel = {col: "{:.2f}" for col in colunas_para_formatar_excel}
styled_df.format(formatter=formatador_excel)

try:
    styled_df.to_excel(nome_arquivo_saida_excel, index=False, engine='openpyxl')
    print(f"\nArquivo Excel visual foi salvo como '{nome_arquivo_saida_excel}'.")

    colunas_a_ignorar_ajuste = ['Descrição', 'Novo Departamento']
    
    workbook = load_workbook(nome_arquivo_saida_excel)
    worksheet = workbook.active
    
    for i, nome_coluna in enumerate(worksheet.columns):
        nome_cabecalho = worksheet.cell(row=1, column=i+1).value
        if nome_cabecalho in colunas_a_ignorar_ajuste:
            continue
            
        max_length = 0
        letra_coluna = get_column_letter(i + 1)
        
        for cell in worksheet[letra_coluna]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        largura_ajustada = (max_length + 2)
        worksheet.column_dimensions[letra_coluna].width = largura_ajustada
        
    workbook.save(nome_arquivo_saida_excel)
    print(f"Largura das colunas no arquivo '{nome_arquivo_saida_excel}' foi ajustada.")
    
except Exception as e:
    print(f"\nOcorreu um erro ao salvar ou ajustar o arquivo Excel: {e}")

# --- GERAÇÃO ARQUIVO.CSV AUTCOM ---
try:
    print(f"Iniciando reestruturação para o arquivo '{nome_arquivo_saida_autcom_csv}'...")
    df_para_autcom = df1.copy()

    df_para_autcom['Cód.Item'] = df_para_autcom['Cód.Item'].apply(
        lambda x: str(int(x)).zfill(7) if pd.notna(x) and x != '' else ''
    )
    
    colunas_para_formatar_csv = colunas_venda + ['Novo Pr.Compra', 'Novo Frete Entrada', 'Preço Real Stihl']
    for col in colunas_para_formatar_csv:
        if col in df_para_autcom.columns:
            df_para_autcom[col] = df_para_autcom[col].apply(
                lambda x: f'{x:.2f}'.replace('.', ',') if isinstance(x, (int, float)) else x
            )
            
    # MUDANÇA: A lista agora só precisa remover a coluna que é exclusiva do Excel
    colunas_para_apagar = ['Preço Real Stihl']
    colunas_existentes_para_apagar = [col for col in colunas_para_apagar if col in df_para_autcom.columns]
    df_para_autcom.drop(columns=colunas_existentes_para_apagar, inplace=True)

    mapeamento_posicoes = {
        'Cód.Item': 0, 'Descrição': 6, 'Referência': 9, 'Novo Pr.Compra': 34,
        'Novo Frete Entrada': 44, 'Novo Pr. Venda 1': 55, 'Novo Pr. Venda 2': 58, 
        'Novo Pr. Venda 3': 61, 'Novo Pr. Venda 4': 64, 'Novo Pr. Venda 5': 67, 
        'Novo Pr. Venda 6': 70, 'Novo Pr. Venda 7': 73,
    }
    df_reestruturado = pd.DataFrame()
    for nome_coluna, nova_posicao in mapeamento_posicoes.items():
        if nome_coluna in df_para_autcom.columns:
            df_reestruturado[nova_posicao] = df_para_autcom[nome_coluna]
    
    todas_as_colunas = range(74)
    df_reestruturado = df_reestruturado.reindex(columns=todas_as_colunas)

    cabecalho_final = [''] * 74
    for nome_coluna, nova_posicao in mapeamento_posicoes.items():
        cabecalho_final[nova_posicao] = nome_coluna
    df_reestruturado.columns = cabecalho_final

    df_reestruturado.to_csv(nome_arquivo_saida_autcom_csv, index=False, sep=';', encoding='latin-1', header=True)
    print(f"Arquivo CSV para Autcom foi salvo como '{nome_arquivo_saida_autcom_csv}'.")

except Exception as e:
    print(f"\nOcorreu um erro ao salvar o arquivo CSV para Autcom: {e}")

print("\nProcesso concluído!")