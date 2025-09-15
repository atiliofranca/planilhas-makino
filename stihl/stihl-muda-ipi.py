import pandas as pd
import tkinter as tk
from tkinter import filedialog
from screeninfo import get_monitors
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- FUNÇÕES AUXILIARES ---

def letra_para_indice(letra):
    letra = letra.upper()
    indice = 0
    for char in letra:
        indice = indice * 26 + (ord(char) - ord('A') + 1)
    return indice - 1

def centralizar_janela_raiz(root):
    try:
        monitor_principal = next(m for m in get_monitors() if m.is_primary)
    except StopIteration:
        monitor_principal = get_monitors()[0]
    
    largura_tela = monitor_principal.width
    altura_tela = monitor_principal.height
    offset_x = monitor_principal.x
    offset_y = monitor_principal.y
    pos_x = offset_x + (largura_tela // 2)
    pos_y = offset_y + (altura_tela // 2)
    root.geometry(f'+{pos_x}+{pos_y}')

# --- CONFIGURAÇÃO ---
nome_arquivo_saida_excel = 'stihl/new-stihl-IPI.xlsx'
nome_arquivo_saida_autcom_csv = 'stihl/new-stihl-IPI-autcom.csv'
COLUNA_DE_BUSCA = 'B'

"""
# --- LISTA DE EXCEÇÕES ---
REFERENCIAS_PARA_IGNORAR = [
    '1127-120-1620', '4112-713-4100', '4119-713-4100', '0000-881-9411',
    '0781-389-3004', '7030-319-0000', '7030-516-0000', '7030-319-0001',
    '7030-516-0002', '0781-389-3012', '7030-319-0002'
]
"""

# --- MAPEAMENTO DAS ABAS E COLUNAS (APENAS PARA IPI) ---
mapeamento_abas = {
    'Lançamentos':                    {'referencia': 'F', 'ipi': 'Q'},
    'MS':                             {'referencia': 'E', 'ipi': 'U'},
    'SABRES CORRENTES PINHÕES LIMAS': {'referencia': 'C', 'ipi': 'J'},
    'ROÇADEIRAS E IMPL':              {'referencia': 'F', 'ipi': 'Q'},
    'CJ.CORTE FS':                    {'referencia': 'C', 'ipi': 'K'},
    'Produtos a Bateria':             {'referencia': 'E', 'ipi': 'S'},
    'OUTRAS MÁQUINAS':                {'referencia': 'E', 'ipi': 'S'},
    'OUTROS':                         {'referencia': 'F', 'ipi': 'P'},
    'PEÇAS':                          {'referencia': 'B', 'ipi': 'I'},
    'ACESSÓRIOS':                     {'referencia': 'C', 'ipi': 'J'},
    'Ferramentas':                    {'referencia': 'B', 'ipi': 'H'},
    'Artigos da Marca':               {'referencia': 'B', 'ipi': 'I'},
    'EPIs':                           {'referencia': 'C', 'ipi': 'K'},
}

# --- SELEÇÃO DE ARQUIVOS COM POP-UP ---
root = tk.Tk()
centralizar_janela_raiz(root)
root.update()
root.withdraw()

print("Por favor, selecione os arquivos de entrada nas janelas pop-up...")
nome_arquivo_1 = filedialog.askopenfilename(
    title="Passo 1 de 2: Selecione o arquivo.csv do Autcom (Lista Base)",
    filetypes=(("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*"))
)
if not nome_arquivo_1: print("Seleção cancelada."); exit()

nome_arquivo_2 = filedialog.askopenfilename(
    title="Passo 2 de 2: Selecione o arquivo.xlsx do fornecedor (Fonte de Dados)",
    filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
)
if not nome_arquivo_2: print("Seleção cancelada."); exit()

print(f"\nArquivo 1 selecionado: {nome_arquivo_1}")
print(f"Arquivo 2 selecionado: {nome_arquivo_2}\n")

# --- LEITURA E PRÉ-PROCESSAMENTO ---
try:
    df1 = pd.read_csv(nome_arquivo_1, sep=';', encoding='latin-1', decimal=',')
    
    print("Lendo todas as abas do arquivo2.xlsx...")
    todas_as_abas_df2 = pd.read_excel(nome_arquivo_2, sheet_name=None, header=None, engine='openpyxl')
    print("Leitura concluída. Consolidando dados para busca rápida...")

    dados_consolidados = {}
    for nome_aba, mapa_colunas in mapeamento_abas.items():
        if nome_aba in todas_as_abas_df2:
            df_aba = todas_as_abas_df2[nome_aba]
            indice_ref = letra_para_indice(mapa_colunas['referencia'])
            indice_ipi = letra_para_indice(mapa_colunas['ipi'])
            for index, linha in df_aba.iterrows():
                referencia = linha.get(indice_ref)
                if pd.notna(referencia) and referencia not in dados_consolidados:
                    # MUDANÇA: Armazena também o nome da aba de origem
                    dados_consolidados[referencia] = {
                        'ipi': linha.get(indice_ipi),
                        'aba_origem': nome_aba 
                    }
    
    print(f"{len(dados_consolidados)} referências únicas consolidadas. Arquivos lidos com sucesso!\n")

except Exception as e:
    print(f"Ocorreu um erro ao ler ou processar os arquivos: {e}"); exit()

# --- PROCESSAMENTO PRINCIPAL ---
indices_modificados = []
indices_ignorados = []
# NOVO: Cria a coluna para armazenar a aba de origem no DataFrame principal
df1['Aba de Origem'] = '' 

for index, linha_df1 in df1.iterrows():
    valor_a_procurar = linha_df1['Referência']
    
    dados_encontrados = dados_consolidados.get(valor_a_procurar)

    # A lógica agora é a mesma para itens ignorados e modificados:
    # Se encontramos, registramos a aba de origem.
    if pd.notna(valor_a_procurar) and dados_encontrados:
        
        # NOVO: Pega o nome da aba e preenche a coluna no df1
        aba_encontrada = dados_encontrados['aba_origem']
        df1.loc[index, 'Aba de Origem'] = aba_encontrada
        
        """
        # Verifica se está na lista de exceções
        if valor_a_procurar in REFERENCIAS_PARA_IGNORAR:
            print(f"Aviso: Referência '{valor_a_procurar}' (linha {index + 2}) está na lista de exceções.")
            indices_ignorados.append(index)
            continue # Pula a modificação do IPI, mas mantém a aba de origem
        """
        # Se não for exceção, atualiza o IPI
        valor_ipi_encontrado = dados_encontrados['ipi']
        print(f"Valor '{valor_a_procurar}' encontrado na aba '{aba_encontrada}'! Atualizando IPI para '{valor_ipi_encontrado}' na linha {index + 2}...")
        df1.loc[index, 'Novo IPI Entrada'] = valor_ipi_encontrado
        indices_modificados.append(index)

# --- PREPARAÇÃO PARA SAÍDA ---
# Cria um novo DataFrame apenas com as colunas desejadas para o Excel
df_saida = df1[['Descrição', 'Referência', 'Novo IPI Entrada', 'Aba de Origem']].copy()
# NOVO: Insere uma coluna D em branco para empurrar a "Aba de Origem" para a coluna E
df_saida.insert(3, '', '')


# --- GERAÇÃO ARQUIVO.XLSX ---
def destacar_celulas(linha):
    colunas_alteradas = ['Novo IPI Entrada']
    if linha.name in indices_ignorados:
        return ['background-color: #FFCDD2'] * len(linha)
    elif linha.name in indices_modificados:
        return ['background-color: yellow' if col in colunas_alteradas else '' for col in linha.index]
    else:
        return [''] * len(linha)

# Usa o novo df_saida para estilizar
styled_df = df_saida.style.apply(destacar_celulas, axis=1)

try:
    styled_df.to_excel(nome_arquivo_saida_excel, index=False, engine='openpyxl')
    print(f"\nArquivo Excel visual foi salvo como '{nome_arquivo_saida_excel}'.")

    # Ajuste da largura das colunas
    workbook = load_workbook(nome_arquivo_saida_excel)
    worksheet = workbook.active
    for i, nome_coluna in enumerate(worksheet.columns):
        max_length = 0
        letra_coluna = get_column_letter(i + 1)
        for cell in worksheet[letra_coluna]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        largura_ajustada = (max_length + 2)
        worksheet.column_dimensions[letra_coluna].width = largura_ajustada
    workbook.save(nome_arquivo_saida_excel)
    print(f"Largura das colunas no arquivo '{nome_arquivo_saida_excel}' foi ajustada.")
    
except Exception as e:
    print(f"\nOcorreu um erro ao salvar ou ajustar o arquivo Excel: {e}")

# --- GERAÇÃO ARQUIVO.CSV AUTCOM ---
# Esta seção não foi alterada e continuará gerando o arquivo com as 3 colunas corretas
try:
    print(f"Iniciando reestruturação para o arquivo '{nome_arquivo_saida_autcom_csv}'...")
    
    mapeamento_posicoes = {
        'Descrição': 6, 
        'Referência': 9, 
        'Novo IPI Entrada': 43
    }
    
    df_reestruturado = pd.DataFrame()
    # Pega os dados do DataFrame principal df1, que contém os dados originais e atualizados
    for nome_coluna, nova_posicao in mapeamento_posicoes.items():
        if nome_coluna in df1.columns:
            df_reestruturado[nova_posicao] = df1[nome_coluna]
    
    todas_as_colunas = range(74)
    df_reestruturado = df_reestruturado.reindex(columns=todas_as_colunas)

    cabecalho_final = [''] * 74
    for nome_coluna, nova_posicao in mapeamento_posicoes.items():
        cabecalho_final[nova_posicao] = nome_coluna
    df_reestruturado.columns = cabecalho_final

    df_reestruturado.to_csv(nome_arquivo_saida_autcom_csv, index=False, sep=';', encoding='latin-1', header=True)
    print(f"Arquivo CSV para Autcom foi salvo como '{nome_arquivo_saida_autcom_csv}'.")

except Exception as e:
    print(f"\nOcorreu um erro ao salvar o arquivo CSV para Autcom: {e}")

print("\nProcesso concluído!")